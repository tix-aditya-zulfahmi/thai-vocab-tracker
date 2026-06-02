#!/usr/bin/env python3
"""
Build / refresh the Thai vocabulary SQLite database from vocabulary-tracker.xlsx.

Design goals
------------
- One portable SQLite file (thai-vocab.db) the future web app can read/write.
- Spaced repetition via Leitner boxes (1..5), tracked PER DIRECTION:
    th2id  = see Thai script, guess Indonesian meaning
    id2th  = see Indonesian meaning, recall the Thai word
- Every guess is logged in `attempts`; per-(word, direction) running state lives
  in `review_state`. The app updates both on each answer.

Idempotent: safe to re-run. It (re)creates schema, refreshes the static content
(vocab / phrases / grammar) from the spreadsheet, and ensures every vocab entry
has a review_state row per direction WITHOUT wiping your accumulated attempts.
"""

import os
import re
import sqlite3

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "vocabulary-tracker.xlsx")
DB = os.path.join(HERE, "thai-vocab.db")

DIRECTIONS = ("th2id", "id2th")

# Leitner box -> how many days until the word is due again after a correct answer.
# Box 1 = due every session (just learned / recently wrong); box 5 = effectively mastered.
LEITNER_INTERVALS = {1: 0, 2: 1, 3: 3, 4: 7, 5: 16}
MAX_BOX = 5


def parse_first_learned(raw):
    """Split the spreadsheet's 'Pertama dipelajari' into (iso_date_or_None, raw_source)."""
    if raw is None:
        return None, None
    s = str(raw).strip()
    m = re.match(r"(\d{4}-\d{2}-\d{2})", s)
    return (m.group(1) if m else None), s


def create_schema(cur):
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS vocab (
            id                   INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_type           TEXT NOT NULL DEFAULT 'word',  -- 'word' | 'phrase'
            thai                 TEXT NOT NULL,
            romanization         TEXT,
            tone                 TEXT,                          -- words only
            meaning              TEXT,                          -- Indonesian
            category             TEXT,
            context              TEXT,                          -- phrases: 'Konteks pakai'
            legacy_frequency     INTEGER DEFAULT 0,             -- original 'Frekuensi' count
            first_learned_date   TEXT,                          -- ISO date if parseable
            first_learned_source TEXT,                          -- raw source string
            notes                TEXT,
            created_at           TEXT DEFAULT (datetime('now')),
            UNIQUE (entry_type, thai, meaning)
        );

        CREATE TABLE IF NOT EXISTS review_state (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            vocab_id       INTEGER NOT NULL REFERENCES vocab(id) ON DELETE CASCADE,
            direction      TEXT NOT NULL,                       -- 'th2id' | 'id2th'
            box            INTEGER NOT NULL DEFAULT 1,          -- Leitner box 1..5
            correct_count  INTEGER NOT NULL DEFAULT 0,
            wrong_count    INTEGER NOT NULL DEFAULT 0,
            current_streak INTEGER NOT NULL DEFAULT 0,          -- consecutive correct
            total_seen     INTEGER NOT NULL DEFAULT 0,
            last_reviewed  TEXT,
            next_due       TEXT,                                -- ISO date the word is due
            is_mastered    INTEGER NOT NULL DEFAULT 0,          -- box 5 + streak >= 3
            UNIQUE (vocab_id, direction)
        );

        CREATE TABLE IF NOT EXISTS attempts (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            vocab_id    INTEGER NOT NULL REFERENCES vocab(id) ON DELETE CASCADE,
            direction   TEXT NOT NULL,
            is_correct  INTEGER NOT NULL,                       -- 1 / 0
            box_before  INTEGER,
            box_after   INTEGER,
            response_ms INTEGER,
            answered_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS grammar_notes (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            topic   TEXT,
            note    TEXT,
            example TEXT
        );

        -- One example sentence per word, plus a word-by-word breakdown (JSON array
        -- of {thai, roman, arti}). Populated by load_examples.py.
        CREATE TABLE IF NOT EXISTS examples (
            vocab_id     INTEGER PRIMARY KEY REFERENCES vocab(id) ON DELETE CASCADE,
            thai         TEXT NOT NULL,
            romanization TEXT,
            meaning      TEXT,            -- Indonesian translation of the sentence
            breakdown    TEXT             -- JSON: [{"thai","roman","arti"}, ...]
        );

        -- Box -> interval (days). The app reads this to schedule next_due. Tunable.
        CREATE TABLE IF NOT EXISTS leitner_intervals (
            box           INTEGER PRIMARY KEY,
            interval_days INTEGER NOT NULL
        );

        CREATE INDEX IF NOT EXISTS idx_attempts_vocab  ON attempts(vocab_id, direction);
        CREATE INDEX IF NOT EXISTS idx_review_due      ON review_state(direction, next_due);
        CREATE INDEX IF NOT EXISTS idx_review_box      ON review_state(direction, box);

        -- Words due now, hardest first (low box / many wrongs surface most often).
        CREATE VIEW IF NOT EXISTS v_due_today AS
        SELECT v.id AS vocab_id, v.thai, v.romanization, v.meaning, v.category,
               r.direction, r.box, r.correct_count, r.wrong_count,
               r.current_streak, r.next_due
        FROM review_state r
        JOIN vocab v ON v.id = r.vocab_id
        WHERE r.is_mastered = 0
          AND (r.next_due IS NULL OR r.next_due <= date('now'))
        ORDER BY r.box ASC, r.wrong_count DESC, v.legacy_frequency DESC;

        -- Per-direction progress summary.
        CREATE VIEW IF NOT EXISTS v_progress AS
        SELECT direction,
               COUNT(*)                              AS total,
               SUM(CASE WHEN is_mastered=1 THEN 1 ELSE 0 END) AS mastered,
               SUM(CASE WHEN box=1 THEN 1 ELSE 0 END)        AS in_box1,
               SUM(correct_count)                    AS total_correct,
               SUM(wrong_count)                      AS total_wrong
        FROM review_state
        GROUP BY direction;
        """
    )
    cur.executemany(
        "INSERT OR REPLACE INTO leitner_intervals(box, interval_days) VALUES (?, ?)",
        list(LEITNER_INTERVALS.items()),
    )


def load_static_content(cur):
    """Refresh vocab / phrases / grammar from the spreadsheet (content only)."""
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # Wipe & reload only the static content tables. review_state / attempts untouched.
    cur.execute("DELETE FROM grammar_notes")

    # --- Kosakata (words) ---
    ws = wb["Kosakata"]
    words = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        thai, roman, tone, meaning, cat, freq, first, notes = (list(row) + [None] * 8)[:8]
        if not thai:
            continue
        d, src = parse_first_learned(first)
        cur.execute(
            """INSERT INTO vocab
               (entry_type, thai, romanization, tone, meaning, category,
                legacy_frequency, first_learned_date, first_learned_source, notes)
               VALUES ('word', ?, ?, ?, ?, ?, ?, ?, ?, ?)
               ON CONFLICT(entry_type, thai, meaning) DO UPDATE SET
                 romanization=excluded.romanization, tone=excluded.tone,
                 category=excluded.category, legacy_frequency=excluded.legacy_frequency,
                 first_learned_date=excluded.first_learned_date,
                 first_learned_source=excluded.first_learned_source, notes=excluded.notes
            """,
            (thai, roman, tone, meaning, cat, int(freq or 0), d, src, notes),
        )
        words += 1

    # --- Frasa (phrases) ---
    ws = wb["Frasa"]
    phrases = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        thai, roman, meaning, context, freq, first, notes = (list(row) + [None] * 7)[:7]
        if not thai:
            continue
        d, src = parse_first_learned(first)
        cur.execute(
            """INSERT INTO vocab
               (entry_type, thai, romanization, meaning, category, context,
                legacy_frequency, first_learned_date, first_learned_source, notes)
               VALUES ('phrase', ?, ?, ?, 'frasa', ?, ?, ?, ?, ?)
               ON CONFLICT(entry_type, thai, meaning) DO UPDATE SET
                 romanization=excluded.romanization, context=excluded.context,
                 legacy_frequency=excluded.legacy_frequency,
                 first_learned_date=excluded.first_learned_date,
                 first_learned_source=excluded.first_learned_source, notes=excluded.notes
            """,
            (thai, roman, meaning, context, int(freq or 0), d, src, notes),
        )
        phrases += 1

    # --- Catatan tata bahasa (grammar) ---
    ws = wb["Catatan tata bahasa"]
    grammar = 0
    for row in list(ws.iter_rows(values_only=True))[1:]:
        topic, note, example = (list(row) + [None] * 3)[:3]
        if not topic and not note:
            continue
        cur.execute(
            "INSERT INTO grammar_notes(topic, note, example) VALUES (?, ?, ?)",
            (topic, note, example),
        )
        grammar += 1

    return words, phrases, grammar


def seed_review_state(cur):
    """Ensure every vocab entry has a review_state row for each direction (box 1)."""
    cur.execute("SELECT id FROM vocab")
    ids = [r[0] for r in cur.fetchall()]
    added = 0
    for vid in ids:
        for direction in DIRECTIONS:
            cur.execute(
                """INSERT OR IGNORE INTO review_state (vocab_id, direction, box)
                   VALUES (?, ?, 1)""",
                (vid, direction),
            )
            added += cur.rowcount
    return added


def main():
    con = sqlite3.connect(DB)
    con.execute("PRAGMA foreign_keys = ON")
    cur = con.cursor()
    create_schema(cur)
    w, p, g = load_static_content(cur)
    new_states = seed_review_state(cur)
    con.commit()

    cur.execute("SELECT COUNT(*) FROM vocab")
    total_vocab = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM review_state")
    total_states = cur.fetchone()[0]
    con.close()

    print(f"DB written: {DB}")
    print(f"  words={w}  phrases={p}  grammar_notes={g}")
    print(f"  vocab rows total={total_vocab}")
    print(f"  review_state rows total={total_states} (new this run={new_states})")


if __name__ == "__main__":
    main()
